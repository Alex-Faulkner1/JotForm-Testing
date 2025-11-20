"""
JotForm Payment Request Test Suite Runner
Executes predefined test cases and generates Excel reports
"""
import asyncio
import json
from datetime import datetime
from pathlib import Path
from playwright.async_api import async_playwright, Page
from utils import make_test_pdf, extract_efs_ref, create_output_folder
from config import JOTFORM_URL, HEADLESS_MODE, WORKFLOW_WAIT_TIME, REDIRECT_TIMEOUT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Import email automation
from gmail_email_handler import get_approval_link_for_test


class TestResult:
    """Store test execution results."""
    def __init__(self, test_id, test_name, form_type, scenario):
        self.test_id = test_id
        self.test_name = test_name
        self.form_type = form_type
        self.scenario = scenario
        self.status = "NOT_RUN"
        self.efs_ref = None
        self.start_time = None
        self.end_time = None
        self.duration = None
        self.expected_outcome = None
        self.actual_outcome = None
        self.error_message = None
        self.screenshots = []
        self.approval_links = []

    def to_dict(self):
        """Convert to dictionary for Excel export."""
        return {
            'Test ID': self.test_id,
            'Test Name': self.test_name,
            'Form Type': self.form_type,
            'Scenario': self.scenario,
            'Status': self.status,
            'EFS Ref': self.efs_ref,
            'Start Time': self.start_time.strftime('%Y-%m-%d %H:%M:%S') if self.start_time else '',
            'End Time': self.end_time.strftime('%Y-%m-%d %H:%M:%S') if self.end_time else '',
            'Duration (s)': f"{self.duration:.2f}" if self.duration else '',
            'Expected Outcome': self.expected_outcome,
            'Actual Outcome': self.actual_outcome,
            'Error Message': self.error_message or '',
            'Screenshots': ', '.join(self.screenshots),
        }


class TestRunner:
    """Main test runner class."""

    def __init__(self, test_suite_file: str, output_folder: str = "test_results",
                 auto_retrieve_links: bool = True, auto_submit_approvals: bool = True,
                 workflow_wait_time: float = 3.0):  # Reduced from 10.0
        """
        Initialize test runner.

        Args:
            test_suite_file: Path to JSON test suite
            output_folder: Folder for test results
            auto_retrieve_links: Auto-retrieve approval links from Gmail
            auto_submit_approvals: Skip manual review for approve actions
            workflow_wait_time: Seconds to wait for JotForm workflows (default 3.0)
        """
        self.test_suite_file = test_suite_file
        self.output_folder = output_folder
        self.auto_retrieve_links = auto_retrieve_links
        self.auto_submit_approvals = auto_submit_approvals
        self.workflow_wait_time = workflow_wait_time
        self.test_cases = []
        self.results = []
        self.page = None
        self.user_email = None

        # Create output folder
        Path(self.output_folder).mkdir(parents=True, exist_ok=True)

    def load_test_suite(self):
        """Load test suite from JSON file."""
        print(f"\n📂 Loading test suite from: {self.test_suite_file}")
        with open(self.test_suite_file, 'r') as f:
            self.test_cases = json.load(f)
        print(f"✅ Loaded {len(self.test_cases)} test cases")

    def get_user_email(self):
        """Get user email for approval workflows."""
        print("\n" + "=" * 60)
        print("Test Suite Configuration")
        print("=" * 60)
        email = input("Enter your email for approvals [mustapha.jobe0001@gmail.com]: ").strip()
        if not email:
            email = "mustapha.jobe0001@gmail.com"
        self.user_email = email
        print(f"✅ Using email: {self.user_email}")
        print(f"🤖 Auto-retrieve links: {'ENABLED' if self.auto_retrieve_links else 'DISABLED'}")
        print(f"⚡ Auto-submit approvals: {'ENABLED' if self.auto_submit_approvals else 'DISABLED'}")
        print(f"⏱️  Workflow wait time: {self.workflow_wait_time}s")
        print()

    def get_approval_link(self, stage_name: str, efs_ref: str) -> str:
        """
        Get approval link - automatically via Gmail or manual input.

        Args:
            stage_name: Approver stage (e.g., "PCM", "RD", "RCM")
            efs_ref: EFS Reference number

        Returns:
            Approval link URL
        """
        print("\n" + "=" * 60)
        print(f"📧 {stage_name} Approval Link Required")
        print("=" * 60)
        print(f"EFS Ref: {efs_ref}")

        if self.auto_retrieve_links:
            print(f"🤖 Attempting automatic retrieval from Gmail...")

            try:
                # AUTOMATIC RETRIEVAL HERE
                link = get_approval_link_for_test(stage_name, efs_ref)

                if link:
                    print(f"\n✅ Link retrieved automatically!")
                    print(f"🔗 {link}")
                    print("=" * 60 + "\n")
                    return link
                else:
                    print(f"\n⚠️  Automatic retrieval failed")
                    print(f"Falling back to manual input...\n")

            except Exception as e:
                print(f"\n⚠️  Error: {e}")
                print(f"Falling back to manual input...\n")

        # Manual fallback (original code)
        print(f"\nCheck your email for the {stage_name} approval email.")
        if stage_name == "PCM":
            print("Expected link format: https://eel.jotform.com/edit/############?eapeo1e")
        elif stage_name == "RD":
            print("Expected link format: https://eel.jotform.com/edit/############?eapet2e")
        print("=" * 60 + "\n")

        link = input(f"Paste {stage_name} link (or 'skip'): ").strip()
        return link if link.lower() != 'skip' else None

    async def wait_for_redirect(self, initial_url: str, timeout: int = REDIRECT_TIMEOUT) -> bool:
        """Wait for page redirect after form submission."""
        await self.page.wait_for_timeout(timeout)
        current_url = self.page.url

        # Check if we're already on a different page (redirect happened)
        if current_url != initial_url:
            print(f"✅ Redirected to: {self.page.url}")
            return True

        # If still on same page, check for success indicators before retrying
        try:
            # Check if "Thank You" or success message is visible (redirect happened but URL didn't update)
            success_indicators = [
                "text=Thank You",
                "text=Fully Approved",
                "text=This Form is Fully Approved",
            ]

            for indicator in success_indicators:
                element = await self.page.query_selector(indicator)
                if element and await element.is_visible():
                    print(f"✅ Success page detected (URL may not have changed)")
                    return True
        except:
            pass

        # Still on form page, try clicking submit again
        print("⚠️  No redirect detected, retrying...")
        try:
            # Check if submit button still exists
            submit_button = await self.page.query_selector("#input_98")
            if submit_button:
                await self.page.click("#input_98")
                await self.page.wait_for_timeout(timeout)

                # Check again if redirect happened
                if self.page.url != initial_url:
                    print(f"✅ Redirected to: {self.page.url}")
                    return True
        except Exception as e:
            # Button may not exist anymore - check if we're on success page
            if self.page.url != initial_url:
                print(f"✅ Redirected to: {self.page.url}")
                return True
            print(f"⚠️  Error clicking submit button: {e}")

        print("❌ Redirect failed")
        return False

    async def select_approver(self):
        """Select approver from ActiveDirectory dropdown."""
        try:
            await self.page.wait_for_timeout(3000)
            frames = self.page.frames

            ad_frames = [(i, f) for i, f in enumerate(frames)
                        if "ActiveDirectoryDropDown.html" in (f.url or "")]

            if not ad_frames:
                print("⚠️  No ActiveDirectory iframe found")
                return

            for frame_idx, frame in ad_frames:
                try:
                    dropdown = await frame.query_selector("#input_ADDropdown")
                    if dropdown:
                        await frame.wait_for_timeout(1000)
                        options = await dropdown.query_selector_all("option")

                        if len(options) > 1:
                            await frame.select_option("#input_ADDropdown",
                                                     value=await options[1].get_attribute("value"))
                            approver_name = await options[1].text_content()
                            print(f"✅ Selected approver: {approver_name}")
                            return
                except:
                    continue

            print("⚠️  Could not select approver automatically")
        except Exception as e:
            print(f"⚠️  Error selecting approver: {e}")

    async def fill_form(self, test_data: dict):
        """Fill the JotForm with test data."""
        print("\n" + "=" * 60)
        print("Filling Form")
        print("=" * 60)

        await self.page.goto(JOTFORM_URL)
        await self.page.wait_for_timeout(2000)

        # Company/Division
        await self.page.select_option("#input_123", label=test_data["company_division"])
        print(f"✓ Company/Division: {test_data['company_division']}")

        # Location Number
        await self.page.fill("#input_124", test_data["location_number"])
        print(f"✓ Location Number: {test_data['location_number']}")

        # Raised By
        await self.page.fill("#input_131", test_data["raised_by"])
        print(f"✓ Raised By: {test_data['raised_by']}")

        # Payment Request Date (auto-filled)
        date_value = await self.page.input_value("#lite_mode_130")
        print(f"✓ Payment Request Date: {date_value}")

        # Payment Request Type - WITH RETRY LOGIC
        max_retries = 2
        for attempt in range(max_retries):
            try:
                # Wait for dropdown to be ready
                await self.page.wait_for_selector("#input_543", state="visible", timeout=10000)
                await self.page.select_option("#input_543", label=test_data["payment_type"])
                print(f"✓ Payment Type: {test_data['payment_type']}")
                break
            except Exception as e:
                if attempt < max_retries - 1:
                    print(f"⚠️  Attempt {attempt + 1} failed: {e}")
                    print("🔄 Refreshing page and retrying...")
                    await self.page.reload()
                    await self.page.wait_for_load_state("networkidle")
                    await self.page.wait_for_timeout(3000)

                    # Re-fill the initial fields
                    await self.page.select_option("#input_123", label=test_data["company_division"])
                    await self.page.fill("#input_124", test_data["location_number"])
                    await self.page.fill("#input_131", test_data["raised_by"])
                else:
                    print(f"❌ Error selecting payment type: {e}")
                    raise

        # Wait longer for form to fully update based on payment type
        # Goods for Resale and Expense Payment have complex conditional fields
        if test_data["payment_type"] in ["Goods for Resale", "Expense Payment"]:
            await self.page.wait_for_timeout(5000)  # 5 seconds for complex forms
        else:
            await self.page.wait_for_timeout(2000)  # 2 seconds for simpler forms

        # Scroll to ensure fields are in viewport
        await self.page.evaluate("window.scrollBy(0, 300)")

        # Description (required for some payment types)
        try:
            await self.page.fill("#input_701", test_data["description"])
            print(f"✓ Description: {test_data['description']}")
        except Exception as e:
            print(f"ℹ️  Description field not available for this payment type")

        # Payee (required for some payment types)
        try:
            await self.page.fill("#input_547", test_data["payee"])
            print(f"✓ Payee: {test_data['payee']}")
        except Exception as e:
            print(f"ℹ️  Payee field not available for this payment type")

        # ============================================================
        # GOODS FOR RESALE / EXPENSE PAYMENT - CORRECT ORDER & SELECTORS!
        # Based on actual form HTML inspection
        # ============================================================

        # STEP 1: Payment in Advance (appears immediately after Payment Type)
        if "payment_in_advance" in test_data:
            try:
                print("\n🔹 Filling Goods for Resale conditional fields...")
                # Wait for form to settle
                await self.page.wait_for_timeout(2000)

                # CORRECT SELECTOR: #label_input_544_0 (Yes) or #label_input_544_1 (No)
                await self.page.wait_for_selector("#label_input_544_0", timeout=15000, state="visible")

                if test_data["payment_in_advance"]:
                    await self.page.click("#label_input_544_0")  # Yes
                    print("✓ Payment in Advance: Yes")
                else:
                    await self.page.click("#label_input_544_1")  # No
                    print("✓ Payment in Advance: No")
            except Exception as e:
                print(f"⚠️  Payment in Advance field error: {e}")

        # STEP 2: How Many Payable Documents (after Payment in Advance)
        if "payable_documents_count" in test_data:
            try:
                await self.page.wait_for_timeout(2000)

                # CORRECT SELECTOR: #input_549 (not #input_849!)
                await self.page.wait_for_selector("#input_549", timeout=15000, state="visible")
                await self.page.select_option("#input_549", label=test_data["payable_documents_count"])
                print(f"✓ Payable Documents Count: {test_data['payable_documents_count']}")

                # Wait for conditional fields to appear after selection
                await self.page.wait_for_timeout(3000)
                print("   Waiting for conditional fields to appear...")
            except Exception as e:
                print(f"⚠️  Payable Documents dropdown error: {e}")

        # STEP 3: Currency GBP (after Payable Documents dropdown)
        if "currency_gbp" in test_data:
            try:
                # CORRECT SELECTOR: #label_input_619_0 (not #label_input_553_0!)
                await self.page.wait_for_selector("#label_input_619_0", timeout=15000, state="visible")

                if test_data["currency_gbp"]:
                    await self.page.click("#label_input_619_0")  # Yes
                    print("✓ Currency GBP: Yes")
                else:
                    await self.page.click("#label_input_619_1")  # No
                    print("✓ Currency GBP: No")
            except Exception as e:
                print(f"⚠️  Currency GBP field error: {e}")

        # ============================================================
        # STEP 4: Payable Document(s) Section
        # Using field order approach from working script
        # ============================================================
        if test_data.get("payable_documents_count") == "1":
            try:
                print("\n🔹 Filling Payable Document(s) section...")
                await self.page.wait_for_timeout(2000)

                # Scroll to ensure fields are visible
                await self.page.evaluate("window.scrollBy(0, 300)")

                # Upload Document 1 - Attach Document
                # CORRECT ID: input_558 (from working script)
                try:
                    pdf_filename = f"invoice_{test_data.get('invoice_number', 'DOC1')}.pdf"
                    make_test_pdf(pdf_filename)

                    # Use the correct ID for Document 1 upload
                    file_input = await self.page.query_selector("#input_558")
                    if file_input:
                        await file_input.set_input_files(pdf_filename)
                        print(f"✓ Uploaded Payable Document: {pdf_filename}")
                    else:
                        print(f"⚠️  Document upload field #input_558 not found")
                except Exception as e:
                    print(f"⚠️  Document upload error: {e}")

                # Get all visible text input fields (not readonly)
                # The working script fills by order: field 3, 4, 5, 6, 7, 8, 9
                text_inputs = await self.page.query_selector_all("input[type='text']:not([readonly])")

                # Filter to only visible and enabled fields
                visible_inputs = []
                for inp in text_inputs:
                    try:
                        is_visible = await inp.is_visible()
                        is_enabled = await inp.is_enabled()
                        if is_visible and is_enabled:
                            visible_inputs.append(inp)
                    except:
                        pass

                print(f"   Found {len(visible_inputs)} visible text input fields")

                # Field 3: Document Number (INV number)
                if len(visible_inputs) > 3 and "invoice_number" in test_data:
                    try:
                        await visible_inputs[3].fill(test_data["invoice_number"])
                        print(f"✓ Document Number: {test_data['invoice_number']}")
                    except Exception as e:
                        print(f"⚠️  Document number error: {e}")

                # Field 4: Document Date
                if len(visible_inputs) > 4:
                    try:
                        today = datetime.now().strftime("%d/%m/%Y")
                        await visible_inputs[4].fill(today)
                        print(f"✓ Document Date: {today}")
                    except Exception as e:
                        print(f"⚠️  Document date error: {e}")

                # Field 5: Goods (£)
                if len(visible_inputs) > 5 and "value" in test_data:
                    try:
                        goods_value = test_data["value"]
                        await visible_inputs[5].fill(goods_value)
                        print(f"✓ Goods (£): {goods_value}")
                    except Exception as e:
                        print(f"⚠️  Goods value error: {e}")

                # Field 6: Carriage (£)
                if len(visible_inputs) > 6:
                    try:
                        await visible_inputs[6].fill("0")
                        print(f"✓ Carriage (£): 0")
                    except Exception as e:
                        print(f"⚠️  Carriage error: {e}")

                # Field 7: VAT (£)
                if len(visible_inputs) > 7 and "value" in test_data:
                    try:
                        # Calculate VAT as 20% of goods value
                        vat_amount = str(int(float(test_data["value"]) * 0.2))
                        await visible_inputs[7].fill(vat_amount)
                        print(f"✓ VAT (£): {vat_amount}")
                    except Exception as e:
                        print(f"⚠️  VAT error: {e}")

                # Field 8: Subtotal (£)
                if len(visible_inputs) > 8 and "value" in test_data:
                    try:
                        # Subtotal = Goods + VAT
                        subtotal = str(int(float(test_data["value"]) * 1.2))
                        await visible_inputs[8].fill(subtotal)
                        print(f"✓ Subtotal (£): {subtotal}")
                    except Exception as e:
                        print(f"⚠️  Subtotal error: {e}")

                # Field 9: Less Settlement Discount (£) - Optional
                if len(visible_inputs) > 9:
                    try:
                        await visible_inputs[9].fill("0")
                        print(f"✓ Less Settlement Discount (£): 0")
                    except Exception as e:
                        pass

                # Total Payable is auto-calculated

            except Exception as e:
                print(f"⚠️  Error in Payable Documents section: {e}")

        # ============================================================
        # STEP 5: Bank Account Details on Invoice
        # ============================================================
        if test_data.get("bank_details_on_invoice", False) and test_data.get("payment_type") == "Goods for Resale":
            try:
                await self.page.wait_for_timeout(2000)
                # Use radio button approach from working script
                # "Are Bank Account Details on Invoice?"
                # Try to find and click the radio button
                try:
                    # Find radio buttons for bank account details
                    bank_yes_radio = await self.page.query_selector("input[type='radio'][value*='Yes'][name*='Are Bank Account Details on Invoice?']")
                    if bank_yes_radio:
                        await bank_yes_radio.click()
                        print("✓ Bank Account Details on Invoice: Yes")
                    else:
                        # Fallback to label approach
                        bank_labels = await self.page.query_selector_all("label")
                        for label in bank_labels:
                            label_text = await label.inner_text()
                            if "bank account details" in label_text.lower() and "yes" in label_text.lower():
                                await label.click()
                                print("✓ Bank Account Details on Invoice: Yes")
                                break
                except Exception as e:
                    print(f"⚠️  Bank Account Details error: {e}")
            except Exception as e:
                print(f"⚠️  Bank Account Details section error: {e}")

        # ============================================================
        # STEP 6: OTHER ATTACHMENTS - Picking Note/Sales Invoice
        # ============================================================
        if test_data.get("payment_type") == "Goods for Resale":
            try:
                print("\n🔹 Uploading Picking Note/Sales Invoice...")
                await self.page.wait_for_timeout(1000)

                # Scroll to OTHER ATTACHMENTS section
                await self.page.evaluate("window.scrollBy(0, 400)")

                # Create picking note PDF
                picking_note_filename = f"picking_note_{test_data.get('invoice_number', 'PN')}.pdf"
                make_test_pdf(picking_note_filename)

                # CORRECT SELECTOR: #input_644 (from HTML inspection)
                picking_note_selectors = [
                    "#input_644",  # Correct ID from HTML
                    "input[type='file'][id='input_644']",
                    "input[type='file'][name*='q644']",
                ]

                uploaded = False
                for selector in picking_note_selectors:
                    try:
                        file_input = await self.page.query_selector(selector)
                        if file_input:
                            await file_input.set_input_files(picking_note_filename)
                            print(f"✓ Uploaded Picking Note: {picking_note_filename}")
                            uploaded = True
                            break
                    except:
                        continue

                if not uploaded:
                    print(f"⚠️  Could not find picking note upload field (tried #input_644)")
            except Exception as e:
                print(f"⚠️  Picking note upload error: {e}")

        # End of Goods for Resale conditional fields
        # ============================================================

        # Has Invoice (may vary by payment type)
        # Skip for Goods for Resale - already handled above in Payable Documents section
        if test_data.get("has_invoice", False) and test_data.get("payment_type") != "Goods for Resale":
            try:
                has_invoice_selector = "#label_input_604_0"
                invoice_radio = await self.page.query_selector(has_invoice_selector)
                if invoice_radio:
                    await self.page.click(has_invoice_selector)
                    print("✓ Has invoice: Yes")
                    await self.page.wait_for_timeout(2000)  # Wait longer for conditional fields

                    # Scroll to invoice section
                    await self.page.evaluate("window.scrollBy(0, 200)")

                    # Upload PDF (different field for Goods for Resale)
                    try:
                        pdf_filename = f"invoice_{test_data['invoice_number']}.pdf"
                        make_test_pdf(pdf_filename)

                        # Try Goods for Resale upload field first
                        if test_data.get("payment_type") == "Goods for Resale":
                            # Field for "Attach Your Picking Note/Sales Invoice"
                            upload_field = await self.page.query_selector("#input_875")
                            if not upload_field:
                                # Fallback to standard field
                                upload_field = await self.page.query_selector("#input_558")
                        else:
                            # Standard invoice upload field
                            upload_field = await self.page.query_selector("#input_558")

                        if upload_field:
                            await upload_field.set_input_files(pdf_filename)
                            print(f"✓ Uploaded PDF: {pdf_filename}")
                        else:
                            print(f"⚠️  Invoice upload field not found")
                    except Exception as e:
                        print(f"⚠️  Could not upload PDF: {e}")

                    # Invoice Number
                    try:
                        invoice_num_field = await self.page.query_selector("#input_562")
                        if invoice_num_field:
                            await self.page.fill("#input_562", test_data["invoice_number"])
                            print(f"✓ Invoice Number: {test_data['invoice_number']}")
                    except Exception as e:
                        print(f"⚠️  Could not fill invoice number: {e}")

                    # Invoice Date
                    try:
                        invoice_date_field = await self.page.query_selector("#lite_mode_566")
                        if invoice_date_field:
                            today = datetime.now().strftime("%d/%m/%Y")
                            await self.page.fill("#lite_mode_566", today)
                            print(f"✓ Invoice Date: {today}")
                    except Exception as e:
                        print(f"⚠️  Could not fill invoice date: {e}")

                    # Bank details on invoice (for Sponsorship/Charitable Donation)
                    if test_data.get("bank_details_on_invoice", False):
                        try:
                            bank_details_selector = "#label_input_607_0"
                            bank_radio = await self.page.query_selector(bank_details_selector)
                            if bank_radio:
                                await self.page.click(bank_details_selector)
                                print("✓ Bank details on invoice: Yes")
                        except Exception as e:
                            print(f"ℹ️  Bank details on invoice option not available")
                else:
                    print("ℹ️  Has invoice radio button not found")
            except Exception as e:
                print(f"⚠️  Could not process invoice section: {e}")

        # Bank Account Details on Invoice (separate question for Goods for Resale)
        if test_data.get("bank_details_on_invoice", False) and test_data.get("payment_type") == "Goods for Resale":
            try:
                await self.page.wait_for_timeout(1000)
                # Check for the separate bank details question
                bank_account_selector = "#label_input_608_0"  # Goods for Resale specific
                bank_radio = await self.page.query_selector(bank_account_selector)
                if bank_radio:
                    await self.page.click(bank_account_selector)
                    print("✓ Bank Account Details on Invoice: Yes")
            except Exception as e:
                print(f"ℹ️  Bank Account Details question not available")

        # Value (different field structure for Goods for Resale)
        try:
            if test_data.get("payment_type") == "Goods for Resale":
                # For Goods for Resale, fill the "Goods (£)" field
                # Scroll to Total Values section
                await self.page.evaluate("window.scrollBy(0, 400)")
                await self.page.wait_for_timeout(1000)

                # Try to find and fill Goods field
                goods_field = await self.page.query_selector("#input_855")
                if goods_field:
                    await self.page.fill("#input_855", test_data["value"])
                    print(f"✓ Goods Value: £{test_data['value']}")
                else:
                    print("⚠️  Goods value field not found")
            else:
                # Standard value field for other payment types
                await self.page.fill("#input_844", test_data["value"])
                print(f"✓ Value: £{test_data['value']}")
        except Exception as e:
            print(f"⚠️  Could not fill value field: {e}")

        # Select approver (may not exist for all payment types)
        try:
            await self.select_approver()
        except Exception as e:
            print(f"⚠️  Could not select approver (may not be needed): {e}")

        # PCM Email
        try:
            pcm_email_field = await self.page.query_selector("#input_195")
            if pcm_email_field:
                await self.page.fill("#input_195", self.user_email)
                print(f"✓ PCM Email: {self.user_email}")
            else:
                print("ℹ️  PCM Email field not found for this payment type")
        except Exception as e:
            print(f"⚠️  Could not fill PCM email: {e}")

        print("=" * 60)
        input("\n⏸️  Review form and press ENTER to submit...")

    async def submit_stage1(self) -> str:
        """Submit Stage 1 and extract EFS reference."""
        print("\n📤 Submitting Stage 1...")
        initial_url = self.page.url
        await self.page.click("#input_98")

        if not await self.wait_for_redirect(initial_url):
            return None

        page_content = await self.page.content()
        efs_ref = extract_efs_ref(page_content)

        if efs_ref:
            print(f"✅ EFS Ref: {efs_ref}")
        else:
            print("❌ Could not extract EFS Ref")

        return efs_ref

    async def check_already_reviewed(self, stage_name: str) -> bool:
        """Check if form has already been reviewed."""
        await self.page.wait_for_timeout(2000)

        # Check for approval buttons first
        approval_selectors = [
            "input[type='radio'][value='Approve']",
            "input[type='radio'][value='Reject']",
            "#label_input_203_0",  # PCM
            "#label_input_249_0",  # RD
        ]

        for selector in approval_selectors:
            try:
                element = await self.page.query_selector(selector)
                if element and await element.is_visible():
                    return False  # Form needs approval
            except:
                continue

        # Check for "already reviewed" banner
        banner_selectors = [
            "text=You have already reviewed this Form",
            "text=No further action is required at this stage",
        ]

        for selector in banner_selectors:
            try:
                element = await self.page.query_selector(selector)
                if element and await element.is_visible():
                    print(f"ℹ️  {stage_name} form already reviewed")
                    return True
            except:
                continue

        return False

    async def process_approval_stage(self, stage: dict, efs_ref: str, test_output_folder: str) -> dict:
        """Process a single approval stage."""
        stage_name = stage["stage"]
        action = stage["action"]

        print(f"\n{'='*60}")
        print(f"Processing {stage_name} Stage")
        print(f"Action: {action}")
        print(f"{'='*60}")

        # Get approval link
        approval_link = self.get_approval_link(stage_name, efs_ref)
        if not approval_link:
            return {"success": False, "reason": "No approval link provided"}

        # Wait for workflows
        print(f"⏳ Waiting {self.workflow_wait_time}s for workflows...")
        await self.page.wait_for_timeout(int(self.workflow_wait_time * 1000))

        # Navigate to approval form
        await self.page.goto(approval_link)
        await self.page.wait_for_load_state("networkidle")
        await self.page.wait_for_timeout(2000)

        # Check if already reviewed
        if await self.check_already_reviewed(stage_name):
            screenshot_path = f"{test_output_folder}/{stage_name.lower()}_already_reviewed.png"
            await self.page.screenshot(path=screenshot_path)
            return {
                "success": True,
                "action": "already_reviewed",
                "screenshot": screenshot_path
            }

        # Find the appropriate radio button based on stage
        selector_map = {
            "PCM": {"approve": "#label_input_203_0", "reject": "#label_input_203_1"},
            "RCM": {"approve": "#label_input_203_0", "reject": "#label_input_203_1"},  # RCM uses same as PCM
            "RD": {"approve": "#label_input_249_0", "reject": "#label_input_249_1"},
        }

        if stage_name not in selector_map:
            return {"success": False, "reason": f"Unknown stage: {stage_name}"}

        approve_selector = selector_map[stage_name]["approve"]
        reject_selector = selector_map[stage_name]["reject"]

        # Click Approve or Reject
        try:
            if action == "Approve":
                await self.page.wait_for_selector(approve_selector, timeout=10000)
                await self.page.click(approve_selector)
                print(f"✓ Selected: Approve")

                # If approving at PCM or RCM, fill RD email field for next stage (if it exists)
                if stage_name in ["PCM", "RCM"]:
                    try:
                        # Wait briefly for form to update after clicking Approve
                        await self.page.wait_for_timeout(1000)

                        # Check if RD email field exists (for 2-stage approvals)
                        rd_email_field = await self.page.query_selector("#input_244")
                        if rd_email_field:
                            is_visible = await rd_email_field.is_visible()
                            if is_visible:
                                await self.page.fill("#input_244", self.user_email)
                                print(f"✓ RD Email for next stage: {self.user_email}")
                    except Exception:
                        # Field doesn't exist or isn't visible - normal for single-stage approvals
                        pass

            elif action == "Reject":
                await self.page.wait_for_selector(reject_selector, timeout=2000)
                await self.page.click(reject_selector)
                print(f"✓ Selected: Reject")

                # Wait for rejection reason field to appear (it appears dynamically after clicking Reject)
                rejection_text = f"Automated test rejection at {stage_name} stage - EFS {efs_ref}"
                filled = False

                # Try to find and fill the rejection reason field
                # The field usually appears within 3 seconds of clicking Reject
                try:
                    # Try different selectors with explicit waits
                    selectors_to_try = [
                        ("textarea[name*='q237']", "name-based"),
                        ("#input_237", "ID-based"),
                        ("textarea[name*='reason' i]", "name with 'reason'"),
                        ("textarea[aria-label*='reason' i]", "aria-label"),
                    ]

                    for selector, selector_type in selectors_to_try:
                        try:
                            # Wait up to 5 seconds for this selector to appear and be visible
                            await self.page.wait_for_selector(selector, timeout=5000, state="visible")

                            # Field is visible, fill it
                            await self.page.fill(selector, rejection_text)
                            print(f"✓ Rejection reason filled ({selector_type}): {rejection_text}")
                            filled = True
                            break
                        except Exception:
                            # This selector didn't work, try next one
                            continue

                    if not filled:
                        # Last resort - try any visible textarea
                        try:
                            textareas = await self.page.query_selector_all("textarea")
                            for textarea in textareas:
                                if await textarea.is_visible():
                                    current_value = await textarea.input_value()
                                    if not current_value or len(current_value.strip()) == 0:
                                        await textarea.fill(rejection_text)
                                        print(f"✓ Rejection reason filled (generic textarea): {rejection_text}")
                                        filled = True
                                        break
                        except:
                            pass

                except Exception as e:
                    print(f"⚠️  Error finding rejection reason field: {e}")

                if not filled:
                    print(f"⚠️  Could not find rejection reason field to auto-fill - please fill manually")
            else:
                return {"success": False, "reason": f"Unknown action: {action}"}
        except Exception as e:
            screenshot_path = f"{test_output_folder}/{stage_name.lower()}_error.png"
            await self.page.screenshot(path=screenshot_path)
            return {
                "success": False,
                "reason": f"Could not click {action} button: {e}",
                "screenshot": screenshot_path
            }

        # Submit - Skip review for approvals if auto_submit is enabled
        if action == "Approve" and self.auto_submit_approvals:
            print(f"⚡ Auto-submitting approval...")
        else:
            # For rejections or when auto_submit is disabled, ask for manual review
            input(f"\n⏸️  Review {stage_name} form and press ENTER to submit...")

        initial_url = self.page.url
        await self.page.click("#input_98")
        print("📤 Submitting...")

        success = await self.wait_for_redirect(initial_url)

        screenshot_path = f"{test_output_folder}/{stage_name.lower()}_{action.lower()}.png"
        await self.page.screenshot(path=screenshot_path)
        print(f"📸 Screenshot: {screenshot_path}")

        return {
            "success": success,
            "action": action,
            "screenshot": screenshot_path
        }

    async def run_test_case(self, test_case: dict) -> TestResult:
        """Execute a single test case."""
        result = TestResult(
            test_case["test_id"],
            test_case["description"],
            test_case["form_type"],
            test_case["scenario"]
        )

        result.start_time = datetime.now()
        result.expected_outcome = test_case["expected_outcome"]

        print("\n" + "🚀" * 30)
        print(f"EXECUTING: {test_case['test_id']}")
        print(f"Form Type: {test_case['form_type']}")
        print(f"Scenario: {test_case['scenario']}")
        print(f"Description: {test_case['description']}")
        print("🚀" * 30)

        try:
            # Fill and submit form
            await self.fill_form(test_case["test_data"])
            efs_ref = await self.submit_stage1()

            if not efs_ref:
                result.status = "FAIL"
                result.error_message = "Failed to submit Stage 1 or extract EFS Ref"
                result.end_time = datetime.now()
                result.duration = (result.end_time - result.start_time).total_seconds()
                return result

            result.efs_ref = efs_ref

            # Create test-specific output folder
            test_output_folder = f"{self.output_folder}/{result.test_id}_{efs_ref}"
            Path(test_output_folder).mkdir(parents=True, exist_ok=True)

            # Take screenshot after Stage 1
            screenshot_path = f"{test_output_folder}/stage1_submitted.png"
            await self.page.screenshot(path=screenshot_path)
            result.screenshots.append(screenshot_path)

            # Process approval workflow
            workflow_results = []
            for stage in test_case["approval_workflow"]:
                stage_result = await self.process_approval_stage(
                    stage, efs_ref, test_output_folder
                )
                workflow_results.append(stage_result)

                if stage_result.get("screenshot"):
                    result.screenshots.append(stage_result["screenshot"])

                if not stage_result["success"]:
                    result.status = "FAIL"
                    result.error_message = stage_result.get("reason", "Unknown error")
                    result.actual_outcome = "Workflow failed"
                    result.end_time = datetime.now()
                    result.duration = (result.end_time - result.start_time).total_seconds()
                    return result

                # If rejected at App 1, workflow ends
                if stage_result.get("action") == "Reject" and stage["stage"] in ["PCM", "RCM"]:
                    result.actual_outcome = "Form DELETED"
                    break

                # If rejected at App 2, should return to App 1
                if stage_result.get("action") == "Reject" and stage["stage"] in ["RD", "COO"]:
                    result.actual_outcome = "Sent back to App 1"

                    # VALIDATE: Check for return-to-PCM email
                    print("\n" + "=" * 60)
                    print("🔍 Validating Return-to-PCM Email")
                    print("=" * 60)
                    print(f"After RD rejection, form should return to PCM...")
                    print(f"Checking for new PCM approval email for {efs_ref}...")

                    try:
                        # Try to get the return-to-PCM approval link
                        return_link = self.get_approval_link("PCM", efs_ref)

                        if return_link and "eapeo1e" in return_link:
                            print(f"✅ VALIDATION PASSED: Return-to-PCM email found!")
                            print(f"   Email contains PCM approval link with eapeo1e suffix")
                            print(f"🔗 {return_link}")
                            result.notes = f"Return email validated: {return_link}"
                        else:
                            print(f"⚠️  VALIDATION WARNING: Could not find return-to-PCM email")
                            print(f"   Expected email with eapeo1e suffix (PCM stage)")
                            result.notes = "Return email not found or invalid"
                    except Exception as e:
                        print(f"⚠️  Error during return email validation: {e}")
                        result.notes = f"Return email validation error: {e}"

                    print("=" * 60)
                    break

            # Determine final outcome
            if all(wr["success"] for wr in workflow_results):
                if all(wr.get("action") == "Approve" for wr in workflow_results):
                    result.actual_outcome = "Form fully approved"
                    result.status = "PASS"
                else:
                    # Outcome was already set during workflow (rejection scenarios)
                    # Check if actual outcome matches expected outcome (flexible matching)
                    expected = result.expected_outcome.lower()
                    actual = result.actual_outcome.lower()

                    # Flexible matching for rejection scenarios
                    if ("deleted" in expected and "deleted" in actual) or \
                       ("sent back" in expected and "sent back" in actual) or \
                       ("return" in expected and "sent back" in actual) or \
                       (expected in actual) or \
                       (actual in expected):
                        result.status = "PASS"
                    else:
                        result.status = "FAIL"
                        result.error_message = f"Expected '{result.expected_outcome}' but got '{result.actual_outcome}'"

            result.end_time = datetime.now()
            result.duration = (result.end_time - result.start_time).total_seconds()

            print(f"\n✅ Test {result.test_id} completed: {result.status}")

        except Exception as e:
            result.status = "ERROR"
            result.error_message = str(e)
            result.end_time = datetime.now()
            result.duration = (result.end_time - result.start_time).total_seconds()
            print(f"\n❌ Test {result.test_id} error: {e}")
            import traceback
            traceback.print_exc()

        return result

    def generate_excel_report(self):
        """Generate Excel report with test results."""
        print("\n📊 Generating Excel report...")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Results"

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        error_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            'Test ID', 'Test Name', 'Form Type', 'Scenario', 'Status',
            'EFS Ref', 'Start Time', 'End Time', 'Duration (s)',
            'Expected Outcome', 'Actual Outcome', 'Error Message'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Data rows
        for row_num, result in enumerate(self.results, 2):
            data = result.to_dict()

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = data.get(header, '')
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = border

                # Color code status
                if header == 'Status':
                    if data['Status'] == 'PASS':
                        cell.fill = pass_fill
                    elif data['Status'] == 'FAIL':
                        cell.fill = fail_fill
                    elif data['Status'] == 'ERROR':
                        cell.fill = error_fill

        # Adjust column widths
        column_widths = {
            'A': 12,  # Test ID
            'B': 40,  # Test Name
            'C': 20,  # Form Type
            'D': 25,  # Scenario
            'E': 12,  # Status
            'F': 15,  # EFS Ref
            'G': 20,  # Start Time
            'H': 20,  # End Time
            'I': 12,  # Duration
            'J': 30,  # Expected
            'K': 30,  # Actual
            'L': 50,  # Error
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws['A1'] = "Test Execution Summary"
        summary_ws['A1'].font = Font(bold=True, size=14)

        total_tests = len(self.results)
        passed = sum(1 for r in self.results if r.status == "PASS")
        failed = sum(1 for r in self.results if r.status == "FAIL")
        errors = sum(1 for r in self.results if r.status == "ERROR")

        summary_data = [
            ["", ""],
            ["Total Tests", total_tests],
            ["Passed", passed],
            ["Failed", failed],
            ["Errors", errors],
            ["Pass Rate", f"{(passed/total_tests*100):.1f}%" if total_tests > 0 else "N/A"],
            ["", ""],
            ["Execution Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]

        for row_num, (label, value) in enumerate(summary_data, 2):
            summary_ws[f'A{row_num}'] = label
            summary_ws[f'B{row_num}'] = value
            if label:
                summary_ws[f'A{row_num}'].font = Font(bold=True)

        summary_ws.column_dimensions['A'].width = 20
        summary_ws.column_dimensions['B'].width = 20

        # Save workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_file = f"{self.output_folder}/test_results_{timestamp}.xlsx"
        wb.save(report_file)

        print(f"✅ Excel report saved: {report_file}")
        return report_file

    async def run_all_tests(self):
        """Execute all test cases in the suite."""
        self.load_test_suite()
        self.get_user_email()

        print(f"\n{'='*60}")
        print(f"Starting test execution: {len(self.test_cases)} tests")
        print(f"{'='*60}\n")

        async with async_playwright() as p:
            browser = await p.firefox.launch(headless=HEADLESS_MODE)
            self.page = await browser.new_page()

            try:
                for i, test_case in enumerate(self.test_cases, 1):
                    print(f"\n{'#'*60}")
                    print(f"TEST {i} of {len(self.test_cases)}")
                    print(f"{'#'*60}")

                    result = await self.run_test_case(test_case)
                    self.results.append(result)

                    # Brief pause between tests
                    if i < len(self.test_cases):
                        print("\n⏸️  Pausing 2 seconds before next test...")
                        await self.page.wait_for_timeout(2000)

            finally:
                print("\n⏸️  Closing browser in 5 seconds...")
                await self.page.wait_for_timeout(5000)
                await browser.close()

        # Generate report
        report_file = self.generate_excel_report()

        # Print summary
        print(f"\n{'='*60}")
        print("TEST EXECUTION COMPLETE")
        print(f"{'='*60}")
        print(f"Total Tests: {len(self.results)}")
        print(f"Passed: {sum(1 for r in self.results if r.status == 'PASS')}")
        print(f"Failed: {sum(1 for r in self.results if r.status == 'FAIL')}")
        print(f"Errors: {sum(1 for r in self.results if r.status == 'ERROR')}")
        print(f"\n📊 Report: {report_file}")
        print(f"📁 Screenshots: {self.output_folder}")
        print(f"{'='*60}\n")


async def main():
    """Main entry point."""
    runner = TestRunner(
        test_suite_file="test_suite.json",
        output_folder="test_results",
        auto_retrieve_links=True,      # Enable automatic email link retrieval
        auto_submit_approvals=True,    # Skip manual review for approvals
        workflow_wait_time=3.0         # Reduced from 10s to 3s
    )
    await runner.run_all_tests()


if __name__ == "__main__":
    print("\n" + "🧪" * 30)
    print("  JotForm Payment Request - Test Suite Runner")
    print("  🤖 WITH AUTOMATED EMAIL LINK RETRIEVAL 🤖")
    print("  Automated Test Execution with Excel Reporting")
    print("🧪" * 30 + "\n")

    asyncio.run(main())