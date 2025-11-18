"""
JotForm Payment Request Test Suite Runner
Executes predefined test cases and generates Excel reports,
with YOPmail integration for automatic approval link retrieval.
"""

import asyncio
import json
import re
import time
from datetime import datetime
from pathlib import Path

from playwright.async_api import async_playwright, Page
from utils import make_test_pdf, extract_efs_ref, create_output_folder
from config import JOTFORM_URL, HEADLESS_MODE, WORKFLOW_WAIT_TIME, REDIRECT_TIMEOUT

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


# ================================================================
# YOPMAIL + SELENIUM (FIREFOX)
# ================================================================
YOPMAIL_BASE_URL = "https://yopmail.com/"


def create_yopmail_driver() -> webdriver.Firefox:
    """
    Create a Firefox WebDriver instance for YOPmail scraping.
    """
    options = FirefoxOptions()
    # Uncomment if you want headless YOPmail scraping:
    # options.add_argument("-headless")
    driver = webdriver.Firefox(options=options)
    driver.set_window_size(1200, 800)
    return driver


def wait_for_iframe(driver, frame_name: str, timeout: int = 15):
    """
    Wait until an iframe with given name/id is available and switch into it.
    """
    WebDriverWait(driver, timeout).until(
        EC.frame_to_be_available_and_switch_to_it((By.NAME, frame_name))
    )


def get_mail_html_for_message(driver, msg_element) -> str:
    """
    Click a message row element inside the inbox iframe and return the
    HTML of the email body loaded into the 'ifmail' iframe.
    """
    try:
        button = msg_element.find_element(By.CSS_SELECTOR, "button.lm")
        button.click()
    except Exception as e:
        print(f"[YOPMAIL] Could not click message row: {e}")
        return ""

    driver.switch_to.default_content()
    try:
        wait_for_iframe(driver, "ifmail", timeout=20)
    except TimeoutException:
        print("[YOPMAIL] Mail iframe 'ifmail' not found after clicking message.")
        return ""

    html = driver.page_source
    driver.switch_to.default_content()
    return html


def extract_stage_jotform_link(html: str, stage_name: str) -> str | None:
    """
    Extract the stage-specific JotForm edit link from email HTML.
    PCM -> eapeo1e
    RD  -> eapet2e
    """
    stage = stage_name.upper()
    if stage == "PCM":
        pattern = r"(https://eel\.jotform\.com/edit/\d+\?eapeo1e)"
    elif stage == "RD":
        pattern = r"(https://eel\.jotform\.com/edit/\d+\?eapet2e)"
    else:
        pattern = r"(https://eel\.jotform\.com/edit/\d+\?[A-Za-z0-9]+)"
    m = re.search(pattern, html)
    return m.group(1) if m else None


def get_yopmail_approval_link(mailbox_name: str, efs_ref: str, stage_name: str) -> str | None:
    """
    Poll the given YOPmail inbox and return the approval link for the
    given EFS Ref and stage, if found in the subject or body.
    """
    driver = create_yopmail_driver()
    max_attempts = 12
    poll_interval = 5

    try:
        msgs = []
        for attempt in range(1, max_attempts + 1):
            url = f"{YOPMAIL_BASE_URL}?login={mailbox_name}"
            print(f"[YOPMAIL] Loading inbox (attempt {attempt}/{max_attempts}): {url}")
            driver.get(url)

            try:
                wait_for_iframe(driver, "ifinbox", timeout=20)
            except TimeoutException:
                print("[YOPMAIL] Inbox iframe 'ifinbox' not found.")
                msgs = []
            else:
                msgs = driver.find_elements(By.CSS_SELECTOR, "div.m[id^='e_']")
                if msgs:
                    print(f"[YOPMAIL] Found {len(msgs)} message(s) in inbox.")
                    break

            if attempt < max_attempts:
                print(f"[YOPMAIL] No messages yet, waiting {poll_interval}s...")
                time.sleep(poll_interval)

        if not msgs:
            print("[YOPMAIL] No emails found matching criteria.")
            return None

        num_msgs = len(msgs)
        for idx in range(num_msgs):
            driver.switch_to.default_content()
            try:
                wait_for_iframe(driver, "ifinbox", timeout=10)
            except TimeoutException:
                print("[YOPMAIL] Lost inbox iframe while iterating messages.")
                break

            current_msgs = driver.find_elements(By.CSS_SELECTOR, "div.m[id^='e_']")
            if idx >= len(current_msgs):
                break

            msg = current_msgs[idx]
            mail_id = msg.get_attribute("id") or f"<no id idx={idx}>"

            subject_text = ""
            try:
                subject_el = msg.find_element(By.CSS_SELECTOR, "div.lms")
                subject_text = subject_el.text or ""
            except Exception:
                pass

            print(
                f"[YOPMAIL] Inspecting message {idx+1}/{num_msgs} "
                f"({mail_id}) - subject: {subject_text!r}"
            )

            html_body = get_mail_html_for_message(driver, msg)
            if not html_body:
                continue

            # Filter by EFS Ref (subject OR body)
            if efs_ref and (efs_ref not in subject_text and efs_ref not in html_body):
                continue

            link = extract_stage_jotform_link(html_body, stage_name)
            if link:
                print(f"[YOPMAIL] Found {stage_name} link: {link}")
                return link

        print("[YOPMAIL] No matching approval link found in inbox.")
        return None

    finally:
        driver.quit()


# ================================================================
# TEST RESULT MODEL
# ================================================================
class TestResult:
    """Store test execution results."""

    def __init__(self, test_id, test_name, form_type, scenario):
        self.test_id = test_id
        self.test_name = test_name
        self.form_type = form_type
        self.scenario = scenario
        self.status = "NOT_RUN"
        self.efs_ref = None
        self.start_time = None
        self.end_time = None
        self.duration = None
        self.expected_outcome = None
        self.actual_outcome = None
        self.error_message = None
        self.screenshots = []
        self.approval_links = []

    def to_dict(self):
        """Convert to dictionary for Excel export."""
        return {
            "Test ID": self.test_id,
            "Test Name": self.test_name,
            "Form Type": self.form_type,
            "Scenario": self.scenario,
            "Status": self.status,
            "EFS Ref": self.efs_ref,
            "Start Time": self.start_time.strftime("%Y-%m-%d %H:%M:%S") if self.start_time else "",
            "End Time": self.end_time.strftime("%Y-%m-%d %H:%M:%S") if self.end_time else "",
            "Duration (s)": f"{self.duration:.2f}" if self.duration else "",
            "Expected Outcome": self.expected_outcome,
            "Actual Outcome": self.actual_outcome,
            "Error Message": self.error_message or "",
            "Screenshots": ", ".join(self.screenshots),
        }


# ================================================================
# MAIN TEST RUNNER
# ================================================================
class TestRunner:
    """Main test runner class."""

    def __init__(self, test_suite_file: str, output_folder: str = "test_results"):
        self.test_suite_file = test_suite_file
        self.output_folder = output_folder
        self.test_cases = []
        self.results = []
        self.page: Page | None = None
        self.user_email: str | None = None
        self.yopmail_mailbox: str | None = None

        Path(self.output_folder).mkdir(parents=True, exist_ok=True)

    # ---------- Setup / Config ----------
    def load_test_suite(self):
        """Load test suite from JSON file."""
        print(f"\n📂 Loading test suite from: {self.test_suite_file}")
        with open(self.test_suite_file, "r") as f:
            self.test_cases = json.load(f)
        print(f"✅ Loaded {len(self.test_cases)} test cases")

    def get_user_email(self):
        """Get user email for approval workflows (and configure YOPmail if used)."""
        print("\n" + "=" * 60)
        print("Test Suite Configuration")
        print("=" * 60)

        email = input("Enter your email for approvals [alex.faulkner@digiblu.com]: ").strip()
        if not email:
            email = "alex.faulkner@digiblu.com"

        self.user_email = email

        if email.lower().endswith("@yopmail.com"):
            self.yopmail_mailbox = email.split("@")[0]
            print(f"✅ Using YOPmail inbox: {self.yopmail_mailbox}@yopmail.com")
        else:
            self.yopmail_mailbox = None

        print(f"✅ Using email: {self.user_email}\n")

    def get_approval_link(self, stage_name: str, efs_ref: str) -> str | None:
        """
        Get the approval link for a given stage and EFS Ref.

        If self.yopmail_mailbox is set (a @yopmail.com address is
        being used), this first tries to automatically fetch the correct
        approval link from that YOPmail inbox. If that fails, it falls
        back to a manual paste prompt.
        """
        if self.yopmail_mailbox:
            print("\n[YOPMAIL] Attempting to auto-fetch approval link from inbox...")
            link = get_yopmail_approval_link(self.yopmail_mailbox, efs_ref, stage_name)
            if link:
                print("[YOPMAIL] Approval link found automatically.\n")
                return link

        # Fallback: manual
        print("\n" + "=" * 60)
        print(f"📧 {stage_name} Approval Link Required")
        print("=" * 60)
        print(f"EFS Ref: {efs_ref}")
        print(f"\nCheck your email for the {stage_name} approval email.")
        if stage_name.upper() == "PCM":
            print("Expected link format: https://eel.jotform.com/edit/############?eapeo1e")
        elif stage_name.upper() == "RD":
            print("Expected link format: https://eel.jotform.com/edit/############?eapet2e")
        print("=" * 60 + "\n")

        link = input(f"Paste {stage_name} link (or 'skip'): ").strip()
        return link if link.lower() != "skip" else None

    # ---------- Helpers ----------
    async def wait_for_redirect(self, initial_url: str, timeout: int = REDIRECT_TIMEOUT) -> bool:
        """Wait for page redirect or visible success message after submission."""
        if not self.page:
            return False

        await self.page.wait_for_timeout(timeout)
        current_url = self.page.url

        if current_url != initial_url:
            print(f"✅ Redirected to: {current_url}")
            return True

        try:
            success_indicators = [
                "text=Thank You",
                "text=Fully Approved",
                "text=This Form is Fully Approved",
            ]
            for indicator in success_indicators:
                if await self.page.is_visible(indicator):
                    print("✅ Success message visible, treating as redirect success")
                    return True
        except Exception:
            pass

        print("⚠️ No redirect or success message detected after timeout")
        return False

    async def fill_form(self, test_case: dict, stage_name: str = "Stage 1"):
        """Fill JotForm based on test case data."""
        if not self.page:
            return

        tc = test_case
        test_id_for_display = (
            tc.get("Test ID")
            or tc.get("ID")
            or tc.get("Test_ID")
            or "UNKNOWN"
        )

        print("\n" + "=" * 60)
        print(f"🧪 Running {stage_name} - Test ID: {test_id_for_display}")
        print(f"Scenario: {tc.get('Scenario', tc.get('scenario', ''))}")
        print("=" * 60)

        # Applicant details
        if stage_name == "Stage 1":
            await self.page.fill("#input_1", tc.get("Applicant Name", "Test User"))
            await self.page.fill("#input_2", tc.get("Applicant Email", self.user_email or ""))

        # Payment type
        payment_type = tc.get("Payment Request Type") or tc.get("payment_request_type")
        if payment_type:
            print(f"🔽 Selecting Payment Request Type: {payment_type}")
            await self.page.select_option("#input_3", label=payment_type)
            await self.page.wait_for_timeout(2000)

        # Dynamic fields
        dynamic_fields = tc.get("Dynamic Fields", {}) or tc.get("dynamic_fields", {})
        for field_id, value in dynamic_fields.items():
            try:
                print(f"✏️  Filling {field_id} with '{value}'")
                if field_id.startswith("dropdown_"):
                    await self.page.select_option(f"#{field_id}", label=value)
                elif field_id.startswith("radio_"):
                    await self.page.check(f"#{field_id}_{value}")
                else:
                    await self.page.fill(f"#{field_id}", str(value))
            except Exception as e:
                print(f"⚠️  Could not fill field {field_id}: {e}")

        # Upload PDF if requested
        if tc.get("Upload PDF") or tc.get("upload_pdf"):
            safe_test_id = (
                tc.get("Test ID")
                or tc.get("ID")
                or tc.get("Test_ID")
                or "UNKNOWN"
            )
            pdf_path = make_test_pdf(safe_test_id, tc.get("Scenario", ""), self.output_folder)
            print(f"📎 Uploading PDF: {pdf_path}")
            try:
                await self.page.set_input_files("#input_4", pdf_path)
            except Exception as e:
                print(f"⚠️  PDF upload failed: {e}")

        # PCM/RD emails
        if stage_name == "Stage 1":
            pcm_email = tc.get("PCM Email", self.user_email or "")
            rd_email = tc.get("RD Email", self.user_email or "")

            try:
                await self.page.fill("#input_95", pcm_email)
                print(f"📧 PCM Email: {pcm_email}")
            except Exception:
                print("ℹ️  PCM Email field not found for this payment type")

            try:
                await self.page.fill("#input_96", rd_email)
                print(f"📧 RD Email: {rd_email}")
            except Exception:
                print("ℹ️  RD Email field not found for this payment type")

        print("=" * 60)
        input("\n⏸️  Review form in browser and press ENTER to submit...")

    async def submit_stage1(self) -> str | None:
        """Submit Stage 1 and extract EFS reference."""
        if not self.page:
            return None

        print("\n📤 Submitting Stage 1...")
        initial_url = self.page.url
        await self.page.click("#input_98")

        if not await self.wait_for_redirect(initial_url):
            return None

        page_content = await self.page.content()
        efs_ref = extract_efs_ref(page_content)

        if efs_ref:
            print(f"✅ EFS Ref: {efs_ref}")
        else:
            print("❌ Could not extract EFS Ref")

        return efs_ref

    async def check_already_reviewed(self) -> bool:
        """Check if form has already been reviewed/approved."""
        if not self.page:
            return False

        print("\n🔍 Checking if form is already reviewed...")
        page_content = await self.page.content()
        if "This Form is Fully Approved" in page_content:
            print("✅ Form is already fully approved - skipping further stages")
            return True
        if "This Form has already been processed" in page_content:
            print("✅ Form has already been processed - skipping further stages")
            return True
        return False

    async def run_approval_stage(self, test_case: dict, efs_ref: str, stage_name: str, link: str) -> str:
        """Run a single approval stage (PCM/RD)."""
        if not self.page:
            return "ERROR"

        print("\n" + "=" * 60)
        print(f"🏛️  {stage_name} Approval Stage")
        print("=" * 60)
        print(f"Link: {link}")
        print("=" * 60)

        await self.page.goto(link, wait_until="domcontentloaded")
        await self.page.wait_for_timeout(3000)

        if await self.check_already_reviewed():
            return "ALREADY_REVIEWED"

        comments_field = test_case.get(f"{stage_name} Comments Field")
        comments_value = test_case.get(f"{stage_name} Comments")
        if comments_field and comments_value:
            try:
                print(f"✏️  Filling {stage_name} comments: {comments_value}")
                await self.page.fill(f"#{comments_field}", comments_value)
            except Exception as e:
                print(f"⚠️  Could not fill {stage_name} comments: {e}")

        decision = test_case.get(f"{stage_name} Decision", "Approve")
        decision_field = test_case.get(f"{stage_name} Decision Field")
        if decision_field:
            try:
                if decision.lower() == "approve":
                    await self.page.check(f"#{decision_field}_yes")
                else:
                    await self.page.check(f"#{decision_field}_no")
                print(f"✅ {stage_name} Decision: {decision}")
            except Exception as e:
                print(f"⚠️  Could not set {stage_name} decision: {e}")

        print("=" * 60)
        input(f"\n⏸️  Review {stage_name} form and press ENTER to submit...")

        try:
            await self.page.click("#input_98")
            print("📤 Submitting...")

            initial_url = self.page.url
            if not await self.wait_for_redirect(initial_url):
                return "NO_REDIRECT"

            page_content = await self.page.content()
            if "This Form is Fully Approved" in page_content:
                print("✅ Form is now fully approved")
                return "APPROVED"
            if "This Form has been sent back" in page_content:
                print("✅ Form has been sent back for revision")
                return "SENT_BACK"
            if "This Form has been rejected" in page_content:
                print("✅ Form has been rejected")
                return "REJECTED"

            print("⚠️  Could not determine final status after submission")
            return "UNKNOWN"

        except Exception as e:
            print(f"❌ Error during {stage_name} submission: {e}")
            return "ERROR"

    async def run_single_test(self, test_case: dict, page: Page) -> TestResult:
        """Run a single test case end-to-end."""
        test_id = (
            test_case.get("Test ID")
            or test_case.get("ID")
            or test_case.get("Test_ID")
            or "UNKNOWN"
        )
        test_name = test_case.get("Test Name") or test_case.get("name") or ""
        form_type = (
            test_case.get("Payment Request Type")
            or test_case.get("payment_request_type")
            or ""
        )
        scenario = test_case.get("Scenario") or test_case.get("scenario") or ""

        result = TestResult(
            test_id=test_id,
            test_name=test_name,
            form_type=form_type,
            scenario=scenario,
        )

        self.page = page
        result.start_time = datetime.now()

        try:
            print("\n" + "🧾" * 30)
            print("  Loading JotForm Payment Request Form")
            print("🧾" * 30 + "\n")

            await page.goto(JOTFORM_URL, wait_until="domcontentloaded")
            await page.wait_for_timeout(3000)

            await self.fill_form(test_case, stage_name="Stage 1")
            efs_ref = await self.submit_stage1()

            if not efs_ref:
                result.status = "FAILED"
                result.error_message = "Could not extract EFS Ref"
                return result

            result.efs_ref = efs_ref

            stages = test_case.get("Approval Stages", []) or test_case.get("approval_stages", [])
            current_status = "PENDING"

            for stage in stages:
                stage_name = stage.get("Stage Name") or stage.get("stage_name") or "UNKNOWN_STAGE"
                print(f"\n🔁 Processing {stage_name} stage...")

                link = self.get_approval_link(stage_name, efs_ref)
                if not link:
                    print(f"⚠️  Skipping {stage_name} stage - no link provided")
                    continue

                stage_result = await self.run_approval_stage(test_case, efs_ref, stage_name, link)
                result.approval_links.append(f"{stage_name}: {link}")

                if stage_result in ["APPROVED", "REJECTED", "SENT_BACK"]:
                    current_status = stage_result
                    if stage_result in ["APPROVED", "REJECTED"]:
                        break

            result.status = current_status
            if current_status == "APPROVED":
                result.actual_outcome = "Form fully approved"
            elif current_status == "REJECTED":
                result.actual_outcome = "Form rejected"
            elif current_status == "SENT_BACK":
                result.actual_outcome = "Form sent back for revision"
            else:
                result.actual_outcome = "Form pending or unknown state"

        except Exception as e:
            result.status = "ERROR"
            result.error_message = str(e)
            print(f"❌ Error in test {result.test_id}: {e}")

        finally:
            result.end_time = datetime.now()
            result.duration = (result.end_time - result.start_time).total_seconds()

        return result

    # ---------- Reporting ----------
    def generate_excel_report(self, filename: str | None = None):
        """Generate Excel report of all test results."""
        print("\n📊 Generating Excel report...")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Results"

        headers = [
            "Test ID",
            "Test Name",
            "Form Type",
            "Scenario",
            "Status",
            "EFS Ref",
            "Start Time",
            "End Time",
            "Duration (s)",
            "Expected Outcome",
            "Actual Outcome",
            "Error Message",
            "Screenshots",
        ]

        ws.append(headers)

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for result in self.results:
            ws.append(
                [
                    result.test_id,
                    result.test_name,
                    result.form_type,
                    result.scenario,
                    result.status,
                    result.efs_ref,
                    result.start_time.strftime("%Y-%m-%d %H:%M:%S") if result.start_time else "",
                    result.end_time.strftime("%Y-%m-%d %H:%M:%S") if result.end_time else "",
                    f"{result.duration:.2f}" if result.duration else "",
                    result.expected_outcome,
                    result.actual_outcome,
                    result.error_message or "",
                    ", ".join(result.screenshots),
                ]
            )

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"JotForm_Test_Results_{timestamp}.xlsx"

        output_path = Path(self.output_folder) / filename
        wb.save(output_path)
        print(f"✅ Excel report generated: {output_path}")

    # ---------- Orchestration ----------
    async def run_all_tests(self):
        """Run all test cases in the suite."""
        self.load_test_suite()
        self.get_user_email()

        _ = create_output_folder(self.output_folder)

        async with async_playwright() as p:
            # Firefox, as requested
            browser = await p.firefox.launch(headless=HEADLESS_MODE)
            page = await browser.new_page()
            self.page = page

            try:
                for test_case in self.test_cases:
                    result = await self.run_single_test(test_case, page)
                    self.results.append(result)

                    print("\n⏸️  Pausing 5 seconds before next test...")
                    await self.page.wait_for_timeout(5000)

            finally:
                print("\n⏸️  Closing browser in 5 seconds...")
                await self.page.wait_for_timeout(5000)
                await browser.close()

        self.generate_excel_report()


# ================================================================
# ENTRY POINT
# ================================================================
async def main():
    runner = TestRunner(
        test_suite_file="test_suite.json",
        output_folder="test_results",
    )
    await runner.run_all_tests()


if __name__ == "__main__":
    print("\n" + "🧪" * 30)
    print("  JotForm Payment Request - Test Suite Runner")
    print("  Automated Test Execution with Excel Reporting")
    print("🧪" * 30 + "\n")

    asyncio.run(main())
